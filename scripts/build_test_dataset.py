"""Build master test dataset from xlsx + docx for symptom-extraction evaluation.

Sources:
  - Assistente pediatrico - Master.xlsx
      * Sheet 'Categorie BOT'  -> canonical symptom catalog (label set)
      * Sheet 'UP enquiries'   -> real user messages with gold symptoms
      * Sheet 'Pediatric agent domains' -> out-of-scope domain reference
  - Domande per testare Agente.docx
      * 8 tables of (expected_symptom(s), test prompt)
      * Section markers from preceding paragraphs:
          'Prompt multipli', 'Prompt molto colloquiali',
          'Mini-set hard mode con routing atteso non ovvio'

Outputs (data/test/):
  - symptom_catalog.json        canonical symptom list (single source of truth)
  - domain_catalog.json         domain list (for out-of-scope routing)
  - test_dataset.jsonl          one record per test case
  - test_dataset.csv            same data, human-readable
  - test_dataset_stats.md       coverage report
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from collections import Counter, defaultdict

import openpyxl
from docx import Document

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Assistente pediatrico - Master.xlsx"
DOCX = ROOT / "Domande per testare Agente.docx"
CATALOG_OUT = ROOT / "data" / "catalog"
EVAL_OUT = ROOT / "data" / "eval"
CATALOG_OUT.mkdir(parents=True, exist_ok=True)
EVAL_OUT.mkdir(parents=True, exist_ok=True)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def clean(s: object) -> str:
    if s is None:
        return ""
    t = str(s).replace("\r", " ").replace("\n", " ").strip()
    # Normalize curly quotes to straight ASCII so labels match downstream.
    t = (
        t.replace("\u2018", "'").replace("\u2019", "'")
        .replace("\u201c", '"').replace("\u201d", '"')
        .replace("\u2013", "-").replace("\u2014", "-")
    )
    t = re.sub(r"\s+", " ", t)
    return t


def split_expected(label_cell: str, canon: set[str] | None = None) -> list[str]:  # noqa: ARG001
    """Deprecated stub kept to avoid import surprises; real logic lives in
    `split_and_canonicalize`."""
    return [label_cell] if label_cell else []


# --------------------------------------------------------------------------- #
# 1. Canonical symptom catalog (Categorie BOT)
# --------------------------------------------------------------------------- #
def load_symptom_catalog(wb) -> tuple[list[dict], list[dict]]:
    """Return (symptoms, all_categories)."""
    ws = wb["Categorie BOT"]
    # Header row is R3
    header = [clean(c.value) for c in ws[3]]
    idx = {h: i for i, h in enumerate(header) if h}

    def col(row, name):
        i = idx.get(name)
        return clean(row[i].value) if i is not None and i < len(row) else ""

    symptoms = []
    categories = []
    for row in ws.iter_rows(min_row=4):
        codice = col(row, "Codice")
        macro = col(row, "Macro Categoria BOT")
        cat_it = col(row, "Categoria (Italian)")
        cat_en = col(row, "Category (EN)")
        if not codice and not cat_it:
            continue
        record = {
            "code": codice,
            "macro_category": macro,
            "label_it": cat_it,
            "label_en": cat_en,
            "triage_depth": col(row, "Profonditα del triage") or col(row, "Profondit\u00e0 del triage"),
            "completion_status": col(row, "Stato completamento"),
            "short_definition": col(row, "Short symptom definition to be embedded in the Agent"),
        }
        categories.append(record)
        if macro.lower().startswith("sintom"):  # 'Sintomi'
            symptoms.append(record)
    return symptoms, categories


def load_domain_catalog(wb) -> list[dict]:
    ws = wb["Pediatric agent domains"]
    domains = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = clean(row[0]) if row else ""
        if not code or not code.upper().startswith("DO"):
            continue
        domains.append({
            "code": code,
            "label_en": clean(row[1]) if len(row) > 1 else "",
            "label_it": clean(row[2]) if len(row) > 2 else "",
            "sensitivity": clean(row[9]) if len(row) > 9 else "",
        })
    return domains


# --------------------------------------------------------------------------- #
# 2. UP enquiries -> test cases
# --------------------------------------------------------------------------- #
def load_up_enquiries(wb) -> list[dict]:
    ws = wb["UP enquiries"]
    cases = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        msg = clean(row[0]) if row else ""
        gold = clean(row[1]) if len(row) > 1 else ""
        verdict = clean(row[2]) if len(row) > 2 else ""
        comment = clean(row[3]) if len(row) > 3 else ""
        retrained_verdict = clean(row[4]) if len(row) > 4 else ""
        if not msg:
            continue
        cases.append({
            "id": f"up_{i-1:03d}",
            "source": "xlsx:UP enquiries",
            "section": "real_user_messages",
            "message": msg,
            "expected_symptoms_raw": gold,
            "verdict_original_run": verdict,
            "verdict_retrained_run": retrained_verdict,
            "notes": comment,
        })
    return cases


# --------------------------------------------------------------------------- #
# 3. DOCX tables -> test cases
# --------------------------------------------------------------------------- #
def section_for_table(idx: int) -> str:
    """Map table index to section label based on the doc's structure.

    Doc structure observed:
        TABLE 0 (81 rows): general single-symptom prompts
        TABLE 1 (61 rows): general single-symptom prompts (cont.)
        Heading: 'Prompt multipli' (multi-symptom)
        TABLE 2 (21 rows): hard-mode multi-symptom (sleep)
        TABLE 3 (17 rows): hard-mode multi-symptom (metabolic)
        TABLE 4 (21 rows): hard-mode multi-symptom (mixed)
        Heading: 'Prompt molto colloquiali'
        TABLE 5 (21 rows): colloquial / non-obvious routing
        Heading: '...'
        TABLE 6 (81 rows): vertigini / capogiro stress test
        TABLE 7 (81 rows): prurito cutaneo diffuso stress test
    """
    return {
        0: "standard_single_symptom_a",
        1: "standard_single_symptom_b",
        2: "multi_symptom_sleep",
        3: "multi_symptom_metabolic",
        4: "multi_symptom_mixed",
        5: "hard_mode_non_obvious_routing",
        6: "stress_vertigini_capogiro",
        7: "stress_prurito_cutaneo",
    }.get(idx, f"table_{idx}")


def load_docx_cases() -> list[dict]:
    doc = Document(DOCX)
    cases = []
    for ti, tbl in enumerate(doc.tables):
        section = section_for_table(ti)
        for ri, row in enumerate(tbl.rows):
            cells = [clean(c.text) for c in row.cells]
            if len(cells) < 2:
                continue
            label_cell, prompt_cell = cells[0], cells[1]
            # Skip header row
            if (
                label_cell.lower().startswith("sintomo")
                or label_cell.lower().startswith("sintomi")
                or label_cell.lower().startswith("routing")
            ):
                continue
            if not prompt_cell:
                continue
            # Strip surrounding quotes from prompts.
            prompt = prompt_cell.strip('"').strip("'").strip()
            cases.append({
                "id": f"doc_t{ti}_r{ri:03d}",
                "source": f"docx:table_{ti}",
                "section": section,
                "message": prompt,
                "expected_symptoms_raw": label_cell,
                "verdict_original_run": "",
                "verdict_retrained_run": "",
                "notes": "",
            })
    return cases


# --------------------------------------------------------------------------- #
# 4. Validate gold labels against the canonical catalog
# --------------------------------------------------------------------------- #
# Aliases observed in the test data -> canonical label in the catalog.
# These bridge minor wording drift between docx prompts and the official list.
ALIASES: dict[str, str] = {
    # parenthesized full forms in catalog vs short form in docx
    "poliuria": "Poliuria (emissione di abbondante quantità di urina)",
    "pollachiuria": "Pollachiuria (necessità di urinare molto spesso ma con piccole quantità)",
    # 'e' vs 'o' connector
    "naso ostruito e che cola": "Naso ostruito o che cola",
    # short forms used in tests
    "sincope o collasso": "Sincope, collasso",
    "sincope": "Sincope, collasso",
    "collasso": "Sincope, collasso",
    "vertigini": "Vertigini, capogiro",
    "capogiro": "Vertigini, capogiro",
    "dolore muscolare": "Dolore muscolare/Scheletrico",
    "dolore muscolare scheletrico": "Dolore muscolare/Scheletrico",
    "dolore muscolare e scheletrico": "Dolore muscolare/Scheletrico",
    "dolore muscolare e/o scheletrico": "Dolore muscolare/Scheletrico",
    "scheletrico": "Dolore muscolare/Scheletrico",
    # slash variants
    "sincope/collasso": "Sincope, collasso",
    "irrequietezza/pianto inconsolabile": "Irrequietezza o pianto inconsolabile",
    "naso ostruito e/o che cola": "Naso ostruito o che cola",
    "prurito o bruciore anale/genitale": "Prurito o bruciore anale o delle aree genitali",
    "prurito/bruciore anale o genitale": "Prurito o bruciore anale o delle aree genitali",
    "arrossamento o gonfiore anale/genitale": "Arrossamento o gonfiore anale o delle aree genitali",
    "arrossamento/gonfiore anale o genitale": "Arrossamento o gonfiore anale o delle aree genitali",
    # 'genitale' shorthand vs full form
    "prurito o bruciore anale o genitale": "Prurito o bruciore anale o delle aree genitali",
    "prurito genitale": "Prurito o bruciore anale o delle aree genitali",
    "prurito o bruciore anale": "Prurito o bruciore anale o delle aree genitali",
    "arrossamento o gonfiore anale o genitale": "Arrossamento o gonfiore anale o delle aree genitali",
    "arrossamento genitale": "Arrossamento o gonfiore anale o delle aree genitali",
    "arrossamento o gonfiore anale": "Arrossamento o gonfiore anale o delle aree genitali",
    # other small drifts
    "ingestione di corpo estraneo": "Ingestione di corpi estranei",
    "russamento": "Russamento nel sonno",
    "irrequietezza": "Irrequietezza o pianto inconsolabile",
    "pianto inconsolabile": "Irrequietezza o pianto inconsolabile",
}


def _resolve(label: str, canon: set[str], canon_lower: dict[str, str]) -> str | None:
    """Return canonical label for `label`, or None if unknown."""
    if label in canon:
        return label
    low = label.lower().strip().strip(".").strip()
    if low in canon_lower:
        return canon_lower[low]
    if low in ALIASES:
        return ALIASES[low]
    return None


def split_and_canonicalize(raw: str, canon: set[str], canon_lower: dict[str, str]) -> tuple[list[str], list[str]]:
    """Return (canonical_labels, unresolved_labels) for a gold cell."""
    if not raw:
        return [], []
    s = raw.strip().strip(".").strip()

    # 1. whole-cell hit
    hit = _resolve(s, canon, canon_lower)
    if hit:
        return [hit], []

    # 2. '+' separator (used for multi-symptom rows in docx)
    if "+" in s:
        parts = [p.strip() for p in s.split("+") if p.strip()]
    else:
        # 3. comma split with re-stitching for catalog labels containing ','
        raw_parts = [p.strip() for p in s.split(",") if p.strip()]
        parts = []
        i = 0
        while i < len(raw_parts):
            if i + 1 < len(raw_parts):
                joined = f"{raw_parts[i]}, {raw_parts[i+1]}"
                if _resolve(joined, canon, canon_lower):
                    parts.append(joined)
                    i += 2
                    continue
            parts.append(raw_parts[i])
            i += 1
        # Final pass: any part still containing '/' that's not itself a
        # canonical label (e.g. 'Sincope/collasso') gets sub-split.
        expanded: list[str] = []
        for p in parts:
            if "/" in p and not _resolve(p, canon, canon_lower):
                expanded.extend(x.strip() for x in p.split("/") if x.strip())
            else:
                expanded.append(p)
        parts = expanded

    canonical: list[str] = []
    unresolved: list[str] = []
    seen: set[str] = set()
    for p in parts:
        if p.lower() in {"fine", "end", "-", "n/a"}:
            continue
        r = _resolve(p, canon, canon_lower)
        if r and r not in seen:
            canonical.append(r)
            seen.add(r)
        elif not r:
            unresolved.append(p)
    return canonical, unresolved


def validate(cases: list[dict], symptoms: list[dict]) -> dict:
    canon = {s["label_it"] for s in symptoms if s["label_it"]}
    canon_lower = {s.lower(): s for s in canon}

    unknown = Counter()
    coverage = Counter()
    for c in cases:
        canonical, unresolved = split_and_canonicalize(
            c["expected_symptoms_raw"], canon, canon_lower
        )
        c["expected_symptoms"] = canonical + unresolved  # for visibility
        c["expected_symptoms_canonical"] = canonical
        c["expected_symptoms_unresolved"] = unresolved
        c["all_labels_in_catalog"] = not unresolved and bool(canonical)
        for lbl in canonical:
            coverage[lbl] += 1
        for lbl in unresolved:
            unknown[lbl] += 1
    return {"unknown_labels": unknown, "label_coverage": coverage, "canon": canon}


# --------------------------------------------------------------------------- #
# 5. Main
# --------------------------------------------------------------------------- #
def main() -> None:
    print(f"Loading {XLSX.name}...")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    symptoms, all_cats = load_symptom_catalog(wb)
    domains = load_domain_catalog(wb)
    up_cases = load_up_enquiries(wb)

    print(f"  symptoms in catalog: {len(symptoms)}")
    print(f"  domains: {len(domains)}")
    print(f"  UP enquiries: {len(up_cases)}")

    print(f"Loading {DOCX.name}...")
    doc_cases = load_docx_cases()
    print(f"  docx test cases: {len(doc_cases)}")

    cases = up_cases + doc_cases
    print(f"Total test cases: {len(cases)}")

    report = validate(cases, symptoms)

    # Write artifacts ------------------------------------------------------- #
    (CATALOG_OUT / "symptom_catalog.json").write_text(
        json.dumps(symptoms, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    with (EVAL_OUT / "test_dataset.jsonl").open("w", encoding="utf-8") as f:
        for c in cases:
            f.write(json.dumps(c, ensure_ascii=False) + "\n")

    with (EVAL_OUT / "test_dataset.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "id", "source", "section", "message",
            "expected_symptoms", "expected_symptoms_raw",
            "all_labels_in_catalog",
            "verdict_original_run", "verdict_retrained_run", "notes",
        ])
        for c in cases:
            w.writerow([
                c["id"], c["source"], c["section"], c["message"],
                "; ".join(c["expected_symptoms_canonical"]),
                c["expected_symptoms_raw"],
                "yes" if c["all_labels_in_catalog"] else "no",
                c["verdict_original_run"], c["verdict_retrained_run"], c["notes"],
            ])

    # Stats ----------------------------------------------------------------- #
    by_section = Counter(c["section"] for c in cases)
    multi = sum(1 for c in cases if len(c["expected_symptoms"]) > 1)
    ok_in_catalog = sum(1 for c in cases if c["all_labels_in_catalog"])

    lines = []
    lines.append("# Master Test Dataset - Coverage Report\n")
    lines.append(f"- canonical symptoms in catalog: **{len(symptoms)}**")
    lines.append(f"- domains: **{len(domains)}**")
    lines.append(f"- total test cases: **{len(cases)}**")
    lines.append(f"  - real user messages (xlsx UP enquiries): {len(up_cases)}")
    lines.append(f"  - synthetic prompts (docx): {len(doc_cases)}")
    lines.append(f"- multi-symptom cases: {multi}")
    lines.append(f"- cases with all gold labels matching catalog: {ok_in_catalog} / {len(cases)}\n")

    lines.append("## By section\n")
    for s, n in sorted(by_section.items(), key=lambda x: -x[1]):
        lines.append(f"- {s}: {n}")

    lines.append("\n## Per-symptom support (gold counts)\n")
    canon = report["canon"]
    cov = report["label_coverage"]
    not_tested = [s for s in canon if s not in cov]
    for label, n in sorted(cov.items(), key=lambda x: -x[1]):
        lines.append(f"- {label}: {n}")

    lines.append(f"\n## Catalog symptoms with **zero** test coverage ({len(not_tested)})\n")
    for s in sorted(not_tested):
        lines.append(f"- {s}")

    if report["unknown_labels"]:
        lines.append("\n## Unknown gold labels (not in catalog)\n")
        for lbl, n in sorted(report["unknown_labels"].items(), key=lambda x: -x[1]):
            lines.append(f"- {lbl}  (x{n})")

    (EVAL_OUT / "test_dataset_stats.md").write_text("\n".join(lines), encoding="utf-8")

    print("\nWrote:")
    for d in (CATALOG_OUT, EVAL_OUT):
        for p in d.iterdir():
            print(f"  {p.relative_to(ROOT)}  ({p.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
